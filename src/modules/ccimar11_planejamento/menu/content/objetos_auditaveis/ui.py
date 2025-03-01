import os
import pandas as pd
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QTableView, QHeaderView,
    QFrame, QGroupBox, QGridLayout, QPushButton, QMessageBox, QFileDialog,
    QScrollArea, QMenu, QTableWidget, QTableWidgetItem
)
from PyQt6.QtGui import QFont, QCursor, QStandardItemModel, QStandardItem
from PyQt6.QtCore import Qt
from .models import ObjetosAuditaveisModel
from .dialogs import MultipicadoresDialog, DetalhesDialog, CriteriosViewDialog
from .persistence import load_multiplicadores, save_multiplicadores, load_objetos_criterios
from .calculations import recalculate_all_objects
from paths import MAT_RELEV_CRIT_PATH

def create_objetos_auditaveis(title_text, database_model):
    """
    Cria um QFrame contendo:
      - Uma tabela principal com colunas: NR, Processos / Objetos Auditáveis,
        Materialidade, Relevância, Criticidade, Total, Tipo de Risco.
      - Tabelas auxiliares para Materialidade, Relevância e Criticidade.
    Retorna o QFrame pronto para ser inserido em qualquer layout PyQt6.
    
    Args:
        title_text (str): Texto do título
        database_model: Modelo de banco de dados
        
    Returns:
        QFrame: Frame contendo a interface de objetos auditáveis
    """
    # Frame principal
    main_frame = QFrame()
    main_layout = QVBoxLayout(main_frame)
    main_layout.setContentsMargins(10, 10, 10, 10)
    main_layout.setSpacing(15)

    # Título e botões em HLayout
    title_layout = QHBoxLayout()
    
    title_label = QLabel("Mapa de Critérios e Pesos")
    title_label.setStyleSheet("font-size: 20px; font-weight: bold; color: #FFFFFF;")
    title_layout.addWidget(title_label)
    
    # Espaçador para empurrar os botões para a direita
    title_layout.addStretch()
    
    # Botões
    btn_export = QPushButton("Exportar")
    btn_export.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
    title_layout.addWidget(btn_export)
    
    btn_import = QPushButton("Importar")
    btn_import.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold;")
    title_layout.addWidget(btn_import)
    
    btn_report = QPushButton("Relatório")
    btn_report.setStyleSheet("background-color: #FF9800; color: white; font-weight: bold;")
    title_layout.addWidget(btn_report)
    
    btn_multiplicadores = QPushButton("Multiplicadores")
    btn_multiplicadores.setStyleSheet("background-color: #9C27B0; color: white; font-weight: bold;")
    title_layout.addWidget(btn_multiplicadores)
    
    main_layout.addLayout(title_layout)
    
    # Carregar multiplicadores
    materialidade_peso, relevancia_peso, criticidade_peso = load_multiplicadores()
    
    # Criar o gerenciador de critérios
    from .criterios_manager import CriteriosManager
    criterios_manager = CriteriosManager()
    
    # Criar o modelo de dados
    model = ObjetosAuditaveisModel(criterios_manager, materialidade_peso=materialidade_peso, relevancia_peso=relevancia_peso, criticidade_peso=criticidade_peso)
    
    # Carregar dados do banco de dados
    data = []
    if database_model:
        # Obter dados do banco de dados
        pass
    
    # Carregar dados do arquivo de critérios
    objetos_criterios = load_objetos_criterios()
    
    # Adicionar objetos ao modelo
    for objeto_id, criterios in objetos_criterios.items():
        # Verificar se criterios é um dicionário
        if not isinstance(criterios, dict):
            continue
            
        if 'valores_calculados' in criterios:
            valores = criterios['valores_calculados']
            row_data = [
                len(data) + 1,  # NR
                objeto_id,  # Descrição
                valores.get('materialidade', 0),  # Materialidade
                valores.get('relevancia', 0),  # Relevância
                valores.get('criticidade', 0),  # Criticidade
                valores.get('total', 0),  # Total
                valores.get('tipo_risco', 'Baixo')  # Tipo de Risco
            ]
            data.append(row_data)
    
    # Carregar dados no modelo
    model.load_data(data)
    
    # Criar a tabela principal
    table_view = QTableView()
    table_view.setModel(model)
    
    # Configurar a tabela
    table_view.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # NR
    table_view.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Descrição
    table_view.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # Materialidade
    table_view.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # Relevância
    table_view.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # Criticidade
    table_view.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)  # Total
    table_view.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)  # Tipo de Risco
    
    # Definir altura das linhas
    table_view.verticalHeader().setDefaultSectionSize(30)
    
    # Adicionar a tabela ao layout
    main_layout.addWidget(table_view)
    
    # Criar tabelas auxiliares
    aux_layout = QHBoxLayout()
    
    # Grupo para Materialidade
    mat_group = QGroupBox(f"Materialidade (x{materialidade_peso})")
    mat_layout = QVBoxLayout(mat_group)
    
    # Grupo para Relevância
    rel_group = QGroupBox(f"Relevância (x{relevancia_peso})")
    rel_layout = QVBoxLayout(rel_group)
    
    # Grupo para Criticidade
    crit_group = QGroupBox(f"Criticidade (x{criticidade_peso})")
    crit_layout = QVBoxLayout(crit_group)
    
    # Função para atualizar as tabelas auxiliares
    def update_auxiliary_tables():
        # Função auxiliar para criar uma tabela
        def create_criteria_table():
            table = QTableView()
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
            table.verticalHeader().setVisible(False)
            table.setEditTriggers(QTableView.EditTrigger.NoEditTriggers)
            table.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
            table.setSelectionMode(QTableView.SelectionMode.SingleSelection)
            return table
        
        # Função auxiliar para criar um layout de critérios
        def create_criteria_layout(criterio_nome):
            layout = QVBoxLayout()
            
            # Título do critério
            title = QLabel(criterio_nome)
            title.setStyleSheet("font-weight: bold;")
            layout.addWidget(title)
            
            # Tabela de critérios
            table = create_criteria_table()
            layout.addWidget(table)
            
            return layout, table
        
        # Função auxiliar para criar um container de scroll
        def create_scroll_container():
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll.setFrameShape(QFrame.Shape.NoFrame)
            
            content = QWidget()
            scroll.setWidget(content)
            
            layout = QVBoxLayout(content)
            layout.setContentsMargins(0, 0, 0, 0)
            layout.setSpacing(10)
            
            return scroll, layout
        
        # Limpar layouts existentes
        for i in reversed(range(mat_layout.count())):
            mat_layout.itemAt(i).widget().setParent(None)
            
        for i in reversed(range(rel_layout.count())):
            rel_layout.itemAt(i).widget().setParent(None)
            
        for i in reversed(range(crit_layout.count())):
            crit_layout.itemAt(i).widget().setParent(None)
        
        # Criar containers de scroll
        mat_scroll, mat_content_layout = create_scroll_container()
        rel_scroll, rel_content_layout = create_scroll_container()
        crit_scroll, crit_content_layout = create_scroll_container()
        
        # Adicionar containers aos layouts
        mat_layout.addWidget(mat_scroll)
        rel_layout.addWidget(rel_scroll)
        crit_layout.addWidget(crit_scroll)
        
        # Obter critérios
        materialidade_criterios = criterios_manager.get_criterios("materialidade")
        relevancia_criterios = criterios_manager.get_criterios("relevancia")
        criticidade_criterios = criterios_manager.get_criterios("criticidade")
        
        # Verificar se os critérios estão no formato esperado
        # Se materialidade_criterios for uma lista de strings, vamos criar um formato compatível
        if materialidade_criterios and len(materialidade_criterios) > 0 and isinstance(materialidade_criterios[0], str):
            # Criar um critério simples para cada string
            materialidade_criterios = [{"nome": criterio, "opcoes": []} for criterio in materialidade_criterios]
        
        if relevancia_criterios and len(relevancia_criterios) > 0 and isinstance(relevancia_criterios[0], str):
            relevancia_criterios = [{"nome": criterio, "opcoes": []} for criterio in relevancia_criterios]
            
        if criticidade_criterios and len(criticidade_criterios) > 0 and isinstance(criticidade_criterios[0], str):
            criticidade_criterios = [{"nome": criterio, "opcoes": []} for criterio in criticidade_criterios]
        
        # Função para criar uma tabela de critérios
        def create_criterios_table(criterios, layout):
            # Título da tabela
            title_label = QLabel("Critérios")
            title_label.setStyleSheet("font-weight: bold; font-size: 14px; margin-top: 10px;")
            layout.addWidget(title_label)
            
            # Criar tabela
            table = QTableWidget()
            table.setColumnCount(3)
            table.setHorizontalHeaderLabels(["Critério", "Descrição", "Pontuação"])
            
            # Configurar a tabela
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
            table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
            table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
            table.setAlternatingRowColors(True)
            table.setStyleSheet("""
                QTableWidget {
                    border: 1px solid #d3d3d3;
                    border-radius: 5px;
                    background-color: #ffffff;
                }
                QHeaderView::section {
                    background-color: #f0f0f0;
                    padding: 4px;
                    font-weight: bold;
                    border: 1px solid #d3d3d3;
                }
                QTableWidget::item:alternate {
                    background-color: #f9f9f9;
                }
            """)
            
            # Variável para contar o número total de linhas
            total_rows = 0
            
            # Para cada critério, adicionar suas opções à tabela
            for criterio in criterios:
                # Verificar se o critério é um dicionário
                if not isinstance(criterio, dict):
                    # Se for uma string, adicionar como critério simples
                    row = table.rowCount()
                    table.insertRow(row)
                    table.setItem(row, 0, QTableWidgetItem(criterio))
                    table.setItem(row, 1, QTableWidgetItem(""))
                    table.setItem(row, 2, QTableWidgetItem("0"))
                    total_rows += 1
                    continue
                    
                criterio_nome = criterio.get("nome", "")
                
                # Para cada opção do critério
                opcoes = criterio.get("opcoes", [])
                if not opcoes:
                    # Se não houver opções, adicionar apenas o critério
                    row = table.rowCount()
                    table.insertRow(row)
                    table.setItem(row, 0, QTableWidgetItem(criterio_nome))
                    table.setItem(row, 1, QTableWidgetItem(""))
                    table.setItem(row, 2, QTableWidgetItem("0"))
                    total_rows += 1
                else:
                    # Para cada opção do critério
                    for opcao in opcoes:
                        # Adicionar uma linha à tabela
                        row = table.rowCount()
                        table.insertRow(row)
                        
                        # Adicionar os itens à linha
                        criterio_item = QTableWidgetItem(criterio_nome)
                        criterio_item.setToolTip(criterio_nome)
                        table.setItem(row, 0, criterio_item)
                        
                        descricao = opcao.get("descricao", "")
                        descricao_item = QTableWidgetItem(descricao)
                        descricao_item.setToolTip(descricao)
                        table.setItem(row, 1, descricao_item)
                        
                        pontuacao = str(opcao.get("pontuacao", 0))
                        pontuacao_item = QTableWidgetItem(pontuacao)
                        pontuacao_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        table.setItem(row, 2, pontuacao_item)
                        
                        # Incrementar o contador de linhas
                        total_rows += 1
            
            # Ajustar a altura da tabela com base no número de linhas
            table_height = min(300, 30 + total_rows * 30)
            table.setMinimumHeight(table_height)
            table.setMaximumHeight(table_height)
            
            # Ajustar o tamanho das linhas
            table.resizeRowsToContents()
            
            # Adicionar a tabela ao layout
            layout.addWidget(table)
            
            return table
        
        # Criar tabelas para cada tipo de critério
        mat_table = create_criterios_table(materialidade_criterios, mat_content_layout)
        rel_table = create_criterios_table(relevancia_criterios, rel_content_layout)
        crit_table = create_criterios_table(criticidade_criterios, crit_content_layout)
        
        # Adicionar espaçador para empurrar as tabelas para o topo
        mat_content_layout.addStretch()
        rel_content_layout.addStretch()
        crit_content_layout.addStretch()
    
    # Adicionar os grupos ao layout
    aux_layout.addWidget(mat_group)
    aux_layout.addWidget(rel_group)
    aux_layout.addWidget(crit_group)
    
    main_layout.addLayout(aux_layout)
    
    # Função para mostrar detalhes de um objeto
    def show_details(index):
        # Obter a linha selecionada
        row = index.row()
        
        # Obter os dados da linha
        row_data = model.get_row_data(row)
        
        if row_data:
            # Abrir o diálogo de detalhes
            dialog = DetalhesDialog(row_data, criterios_manager, model, row, main_frame)
            dialog.exec()
    
    # Conectar o clique duplo na tabela à função de mostrar detalhes
    table_view.doubleClicked.connect(show_details)
    
    # Adicionar menu de contexto à tabela
    table_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
    
    # Função para exportar para Excel
    def export_to_excel():
        # Abrir diálogo para selecionar o arquivo
        file_path, _ = QFileDialog.getSaveFileName(
            main_frame,
            "Exportar para Excel",
            "",
            "Arquivos Excel (*.xlsx)"
        )
        
        if file_path:
            try:
                # Criar um novo workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Objetos Auditáveis"
                
                # Adicionar cabeçalhos
                headers = [
                    "NR",
                    "Processos / Objetos Auditáveis",
                    f"Materialidade (x{materialidade_peso})",
                    f"Relevância (x{relevancia_peso})",
                    f"Criticidade (x{criticidade_peso})",
                    "Total",
                    "Tipo de Risco"
                ]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                
                # Adicionar dados
                for row in range(model.rowCount()):
                    for col in range(model.columnCount()):
                        cell = ws.cell(row=row+2, column=col+1)
                        cell.value = model.data(model.index(row, col))
                
                # Ajustar largura das colunas
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 15
                
                # Salvar o arquivo
                wb.save(file_path)
                
                QMessageBox.information(main_frame, "Exportação Concluída", "Os dados foram exportados com sucesso!")
            except Exception as e:
                QMessageBox.critical(main_frame, "Erro na Exportação", f"Ocorreu um erro ao exportar os dados: {str(e)}")
    
    # Função para importar do Excel
    def import_from_excel():
        # Abrir diálogo para selecionar o arquivo
        file_path, _ = QFileDialog.getOpenFileName(
            main_frame,
            "Importar do Excel",
            "",
            "Arquivos Excel (*.xlsx)"
        )
        
        if file_path:
            try:
                # Ler o arquivo Excel
                df = pd.read_excel(file_path)
                
                # Limpar o modelo atual
                model.clear()
                
                # Converter o DataFrame para uma lista de listas
                data = df.values.tolist()
                
                # Carregar os dados no modelo
                model.load_data(data)
                
                QMessageBox.information(main_frame, "Importação Concluída", "Os dados foram importados com sucesso!")
            except Exception as e:
                QMessageBox.critical(main_frame, "Erro na Importação", f"Ocorreu um erro ao importar os dados: {str(e)}")
    
    # Função para gerar relatório
    def generate_report():
        pass
    
    # Função para editar multiplicadores
    def edit_multiplicadores():
        # Obter os valores atuais dos multiplicadores
        nonlocal materialidade_peso, relevancia_peso, criticidade_peso
        
        # Abrir o diálogo de multiplicadores
        dialog = MultipicadoresDialog(main_frame, materialidade_peso, relevancia_peso, criticidade_peso)
        
        # Se o diálogo for aceito
        if dialog.exec():
            # Obter os novos valores
            materialidade_peso, relevancia_peso, criticidade_peso = dialog.get_multiplicadores()
            
            # Salvar os novos valores
            save_multiplicadores(materialidade_peso, relevancia_peso, criticidade_peso)
            
            # Atualizar o modelo
            model.update_multiplicadores(materialidade_peso, relevancia_peso, criticidade_peso)
            
            # Atualizar os títulos dos grupos
            mat_group.setTitle(f"Materialidade (x{materialidade_peso})")
            rel_group.setTitle(f"Relevância (x{relevancia_peso})")
            crit_group.setTitle(f"Criticidade (x{criticidade_peso})")
            
            # Forçar atualização da view
            table_view.reset()
            
            # Atualizar tabelas auxiliares
            update_auxiliary_tables()

    # Atualizar tabelas auxiliares
    update_auxiliary_tables()

    # Conectar botões às funções
    btn_export.clicked.connect(export_to_excel)
    btn_import.clicked.connect(import_from_excel)
    btn_report.clicked.connect(generate_report)
    btn_multiplicadores.clicked.connect(edit_multiplicadores)

    return main_frame 